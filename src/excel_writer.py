# ============================================================================
# src/excel_writer.py - PROYECTO_DIOT
# Generador de Excel multi-pestaña con formato
# ============================================================================

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# Estilo del header
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def write_excel(reports: dict, output_path: Path, sheet_order: list = None):
    """
    Escribe múltiples DataFrames como pestañas de un archivo Excel.

    Args:
        reports: Dict {nombre_pestaña: DataFrame}
        output_path: Ruta del archivo Excel de salida.
        sheet_order: Lista de nombres de pestaña en orden deseado (opcional).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    LARGE_THRESHOLD = 5000  # Filas para considerar "grande"

    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # Ordenar pestañas
    if sheet_order:
        ordered_names = [name for name in sheet_order if name in reports]
        # Agregar las que no están en el orden explícito
        ordered_names += [name for name in reports if name not in ordered_names]
    else:
        ordered_names = list(reports.keys())

    for sheet_name in ordered_names:
        df = reports[sheet_name]
        if df is None or df.empty:
            logger.warning(f"⚠️ Pestaña '{sheet_name}' vacía, se omite")
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
        is_large = len(df) > LARGE_THRESHOLD

        # Escribir headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Escribir datos
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Solo aplicar formato en hojas pequeñas
                if not is_large:
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical='center')

        # Auto-ajustar ancho de columnas
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            # Revisar las primeras 100 filas para estimar ancho
            for row in df.head(100).itertuples(index=False):
                val = row[col_idx - 1]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

        # Congelar primera fila
        ws.freeze_panes = 'A2'
        # Filtros
        ws.auto_filter.ref = ws.dimensions

        logger.info(f"📊 Pestaña '{sheet_name}': {len(df):,} registros, {len(df.columns)} columnas")

    if not wb.sheetnames:
        logger.error("❌ No hay pestañas para escribir")
        return None

    wb.save(output_path)
    logger.info(f"💾 Excel guardado: {output_path}")
    return output_path
